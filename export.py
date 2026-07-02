"""Excel export — generates .xlsx attendance reports."""

import os
import calendar
from datetime import datetime, timedelta, date
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
import database_supabase as db
import i18n
from analytics import get_worker_status


# ── Style constants ──────────────────────────────────────────────────────
_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="center", wrap_text=False)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_ON_TIME_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_LATE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_ABSENT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

_LINK_FONT = Font(name="Calibri", color="0563C1", underline="single", size=10)

# Shared alignment objects (reused across cells — creating one per cell is slow on big sheets).
_ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(vertical="center", horizontal="left")

# Russian weekday abbreviations, Monday-first (matches datetime.weekday()).
_RU_WEEKDAYS = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
_RU_MONTHS = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]
_WEEKEND_HEADER_FILL = PatternFill(start_color="7A4F9E", end_color="7A4F9E", fill_type="solid")


def _month_label(year: int, month: int) -> str:
    """e.g. (2026, 6) -> 'Июнь 2026'. Safe as an Excel sheet name."""
    return f"{_RU_MONTHS[month - 1]} {year}"


def _local_dt(ts_str: str, tz: ZoneInfo) -> datetime:
    """Parse an ISO timestamp and convert to the local timezone."""
    ts = datetime.fromisoformat(ts_str)
    if ts.tzinfo is None:
        ts = ts.replace(tzinfo=ZoneInfo("UTC"))
    return ts.astimezone(tz)


def _style_header(ws, col_count: int) -> None:
    """Apply styling to the first (header) row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _auto_width(ws) -> None:
    """Set column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _get_photo_url(file_id: str) -> str:
    """Generate a Telegram Bot API file URL.
    Note: actual download requires a getFile call, but we store
    the file_id so the admin can retrieve it via the bot."""
    return f"https://api.telegram.org/bot{config.BOT_TOKEN}/getFile?file_id={file_id}"


def _get_maps_url(lat: float, lon: float) -> str:
    return f"https://www.google.com/maps?q={lat},{lon}"


def _status_fill(status: str) -> PatternFill:
    if "On Time" in status:
        return _ON_TIME_FILL
    elif "Late" in status:
        return _LATE_FILL
    else:
        return _ABSENT_FILL


def _write_detail_sheet(ws, checkins: list[dict], tz: ZoneInfo) -> None:
    """Write the per-check-in detail table (one row per check-in) onto ``ws``."""
    headers = [
        "#", "Дата", "Время", "Группа", "Юзернейм",
        "Имя", "Фамилия", "Широта", "Долгота",
        "Карта Google", "Тип медиа", "Ссылка",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    for idx, c in enumerate(checkins, start=1):
        ts = _local_dt(c["timestamp"], tz)

        maps_url = ""
        if c.get("latitude") and c.get("longitude"):
            maps_url = _get_maps_url(c["latitude"], c["longitude"])

        media_url = ""
        if c.get("media_file_id"):
            media_url = _get_photo_url(c["media_file_id"])

        row_num = idx + 1  # 1-indexed, skip header
        ws.append([
            idx,
            ts.strftime("%Y-%m-%d"),
            ts.strftime("%H:%M"),
            c.get("group_name", ""),
            c.get("username", ""),
            c.get("first_name", ""),
            c.get("last_name", ""),
            c.get("latitude", ""),
            c.get("longitude", ""),
            maps_url,
            c.get("media_type", "photo"),
            media_url,
        ])

        # Make links clickable
        if maps_url:
            cell = ws.cell(row=row_num, column=10)
            cell.hyperlink = maps_url
            cell.font = _LINK_FONT
            cell.value = "📍 Открыть карту"
        if media_url:
            cell = ws.cell(row=row_num, column=12)
            cell.hyperlink = media_url
            cell.font = _LINK_FONT
            cell.value = "🔗 Посмотреть"

    # Fixed column widths — no per-cell borders/alignment or full-sheet width scan,
    # so this stays fast even with tens of thousands of rows (Excel gridlines remain).
    for col_letter, width in {
        "A": 5, "B": 12, "C": 8, "D": 24, "E": 16, "F": 14,
        "G": 14, "H": 11, "I": 11, "J": 16, "K": 12, "L": 14,
    }.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"  # Freeze header row


def generate_export(checkins: list[dict], title: str = "attendance") -> str:
    """Create an .xlsx file and return its path.

    Parameters
    ----------
    checkins : list[dict]
        Rows from database.get_checkins_for_*
    title : str
        Used in the filename.

    Returns
    -------
    str  – absolute path to the generated file.
    """
    os.makedirs(config.EXPORTS_DIR, exist_ok=True)
    tz = ZoneInfo(config.TIMEZONE)

    # Only export file submissions (exclude location-only entries)
    checkins = [c for c in checkins if c.get("media_file_id")]

    wb = Workbook()

    # ── Sheet 1: Detailed Check-ins ──────────────────────────────────
    ws1 = wb.active
    ws1.title = i18n.EXCEL_SHEET_CHECKINS
    _write_detail_sheet(ws1, checkins, tz)

    # ── Sheet 2: Daily Summary ───────────────────────────────────────
    ws2 = wb.create_sheet(i18n.EXCEL_SHEET_SUMMARY)

    headers2 = [
        "Дата", "Юзернейм", "ФИО", "Группа",
        "Кол-во чекинов", "Первый чекин", "Последний чекин", "Статус",
    ]
    ws2.append(headers2)
    _style_header(ws2, len(headers2))

    # Group checkins by date to compute summary
    dates = sorted(set(c["date"] for c in checkins)) if checkins else []
    row_idx = 2
    all_workers = db.get_active_workers()

    for d in dates:
        summary = db.get_daily_summary(d)
        present_ids = {r["user_id"] for r in summary}

        for r in summary:
            status = get_worker_status(r["first_checkin"])
            first_t = datetime.fromisoformat(r["first_checkin"])
            last_t = datetime.fromisoformat(r["last_checkin"])
            if first_t.tzinfo is None: first_t = first_t.replace(tzinfo=ZoneInfo("UTC"))
            if last_t.tzinfo is None: last_t = last_t.replace(tzinfo=ZoneInfo("UTC"))
            
            first_t = first_t.astimezone(tz)
            last_t = last_t.astimezone(tz)

            ws2.append([
                d,
                r.get("username", ""),
                f"{r.get('first_name', '')} {r.get('last_name', '')}".strip(),
                r.get("group_name", ""),
                r["checkin_count"],
                first_t.strftime("%H:%M"),
                last_t.strftime("%H:%M"),
                status,
            ])
            # Color the status cell
            status_cell = ws2.cell(row=row_idx, column=8)
            status_cell.fill = _status_fill(status)
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=row_idx, column=col).border = _THIN_BORDER
                ws2.cell(row=row_idx, column=col).alignment = _CELL_ALIGN
            row_idx += 1

        # Absent workers
        for w in all_workers:
            if w["user_id"] not in present_ids:
                ws2.append([
                    d,
                    w.get("username", ""),
                    f"{w.get('first_name', '')} {w.get('last_name', '')}".strip(),
                    "—",
                    0,
                    "—",
                    "—",
                    "❌ Отсутствует",
                ])
                status_cell = ws2.cell(row=row_idx, column=8)
                status_cell.fill = _ABSENT_FILL
                for col in range(1, len(headers2) + 1):
                    ws2.cell(row=row_idx, column=col).border = _THIN_BORDER
                    ws2.cell(row=row_idx, column=col).alignment = _CELL_ALIGN
                row_idx += 1

    _auto_width(ws2)
    ws2.freeze_panes = "A2"

    # ── Save ─────────────────────────────────────────────────────────
    now_str = datetime.now(tz).strftime("%Y%m%d_%H%M%S")
    filename = f"{title}_{now_str}.xlsx"
    filepath = os.path.join(config.EXPORTS_DIR, filename)
    wb.save(filepath)
    return filepath


def _write_matrix_sheet(ws, title_text: str, dates: list[str], checkins: list[dict],
                        all_workers: list[dict], last_groups: dict, tz: ZoneInfo) -> None:
    """Write a per-day attendance matrix onto ``ws``.

    One row per worker, one column per date in ``dates``. Each cell lists *all*
    the distinct check-in times that worker sent that day (green), or "—" if
    absent (red). A trailing "Дней" column shows days present / total.
    Works for any span — a week, a month, or a single month of all-time.
    """
    # user_id -> { date -> {"HH:MM", ...} } (unique minutes = distinct send times)
    day_times: dict[int, dict[str, set]] = {}
    for c in checkins:
        ts = _local_dt(c["timestamp"], tz)
        day_times.setdefault(c["user_id"], {}).setdefault(c["date"], set()).add(ts.strftime("%H:%M"))

    def _grp(uid: int) -> str:
        return last_groups.get(uid, "Новички / Новые")

    workers = sorted(
        all_workers,
        key=lambda w: (_grp(w["user_id"]).lower(),
                       (w.get("first_name") or "").lower(),
                       (w.get("last_name") or "").lower()),
    )

    n_days = len(dates)
    total_cols = 4 + n_days + 1  # #, ФИО, Юзернейм, Группа, [days...], Дней

    # Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = _HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row (weekday + dd.mm per day; weekends tinted)
    headers = ["#", i18n.COL_FULL_NAME, i18n.COL_USERNAME, i18n.COL_GROUP]
    for d in dates:
        dt = datetime.strptime(d, "%Y-%m-%d")
        headers.append(f"{_RU_WEEKDAYS[dt.weekday()]}\n{dt.strftime('%d.%m')}")
    headers.append(i18n.COL_DAYS)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
        is_weekend = 5 <= col <= 4 + n_days and datetime.strptime(dates[col - 5], "%Y-%m-%d").weekday() >= 5
        cell.fill = _WEEKEND_HEADER_FILL if is_weekend else _HEADER_FILL
    ws.row_dimensions[2].height = 32

    # Data rows
    row = 3
    for idx, w in enumerate(workers, start=1):
        uid = w["user_id"]
        name = f"{w.get('first_name', '')} {w.get('last_name', '')}".strip() or "—"
        uname = f"@{w['username']}" if w.get("username") else ""
        wmap = day_times.get(uid, {})

        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=uname)
        ws.cell(row=row, column=4, value=_grp(uid))

        present_days = 0
        max_lines = 1
        for i, d in enumerate(dates):
            cell = ws.cell(row=row, column=5 + i)
            times = sorted(wmap.get(d, set()))
            if times:
                present_days += 1
                cell.value = "\n".join(times)
                cell.fill = _ON_TIME_FILL
                max_lines = max(max_lines, len(times))
            else:
                cell.value = "—"
                cell.fill = _ABSENT_FILL
            cell.alignment = _ALIGN_CENTER_WRAP

        cnt = ws.cell(row=row, column=total_cols, value=f"{present_days}/{n_days}")
        cnt.font = Font(bold=True)
        cnt.alignment = _ALIGN_CENTER
        if present_days == 0:
            cnt.fill = _ABSENT_FILL
        elif present_days == n_days:
            cnt.fill = _ON_TIME_FILL
        else:
            cnt.fill = _LATE_FILL

        for col in range(1, total_cols + 1):
            c = ws.cell(row=row, column=col)
            c.border = _THIN_BORDER
            if col in (2, 3, 4):
                c.alignment = _ALIGN_LEFT
            elif col == 1:
                c.alignment = _ALIGN_CENTER

        ws.row_dimensions[row].height = max(16, max_lines * 14)
        row += 1

    # Column widths (manual — content has newlines, so _auto_width would over-size)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22
    for i in range(n_days):
        ws.column_dimensions[get_column_letter(5 + i)].width = 9
    ws.column_dimensions[get_column_letter(total_cols)].width = 8

    ws.freeze_panes = "E3"  # keep title, header & identity columns visible
    if row > 3:
        ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}{row - 1}"


def generate_weekly_export(end_date: str, title: str = "weekly") -> str:
    """Create a weekly attendance .xlsx and return its path.

    The workbook has two sheets:

    * **Недельный обзор** – a 7-day matrix. One row per worker; one column per
      day of the week. Each cell lists *all* the times that worker checked in
      that day (green), or "—" if they were absent (red). A trailing "Дней"
      column shows how many of the 7 days they were present.
    * **Все чекины** – the detailed, one-row-per-check-in table for the week.

    Parameters
    ----------
    end_date : str
        Last day of the 7-day window, ``YYYY-MM-DD``.
    title : str
        Used in the filename.
    """
    os.makedirs(config.EXPORTS_DIR, exist_ok=True)
    tz = ZoneInfo(config.TIMEZONE)

    end = datetime.strptime(end_date, "%Y-%m-%d")
    start = end - timedelta(days=6)
    start_str = start.strftime("%Y-%m-%d")
    week_dates = [(start + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]

    # File submissions only (exclude location-only / text entries).
    checkins = [c for c in db.get_checkins_for_range(start_str, end_date) if c.get("media_file_id")]
    all_workers = db.get_active_workers()
    last_groups = db.get_workers_last_groups()

    wb = Workbook()
    ws = wb.active
    ws.title = i18n.EXCEL_SHEET_WEEKLY
    _write_matrix_sheet(
        ws, f"{i18n.WEEKLY_EXCEL_TITLE}:  {start_str} → {end_date}",
        week_dates, checkins, all_workers, last_groups, tz,
    )

    # ── Sheet 2: Detailed check-ins for the week ─────────────────────
    ws2 = wb.create_sheet(i18n.EXCEL_SHEET_CHECKINS)
    _write_detail_sheet(ws2, checkins, tz)

    # ── Save ─────────────────────────────────────────────────────────
    now_str = datetime.now(tz).strftime("%Y%m%d_%H%M%S")
    filename = f"{title}_{start_str}_{end_date}_{now_str}.xlsx"
    filepath = os.path.join(config.EXPORTS_DIR, filename)
    wb.save(filepath)
    return filepath


def _month_day_list(year: int, month: int, today: date) -> list[str]:
    """All 'YYYY-MM-DD' days in a calendar month, capped at today for the current month."""
    first = date(year, month, 1)
    last = date(year, month, calendar.monthrange(year, month)[1])
    if (year, month) == (today.year, today.month):
        last = min(last, today)  # don't render future days as absent
    return [(first + timedelta(days=i)).strftime("%Y-%m-%d")
            for i in range((last - first).days + 1)]


def generate_monthly_export(month: str = None, title: str = "monthly") -> str:
    """Create a monthly attendance .xlsx (same matrix design as the weekly one).

    Parameters
    ----------
    month : str | None
        Target month as ``YYYY-MM``. Defaults to the current calendar month.
    """
    os.makedirs(config.EXPORTS_DIR, exist_ok=True)
    tz = ZoneInfo(config.TIMEZONE)
    today = datetime.now(tz).date()

    if month:
        year, mon = int(month[:4]), int(month[5:7])
    else:
        year, mon = today.year, today.month

    dates = _month_day_list(year, mon, today)
    checkins = [c for c in db.get_checkins_for_range(dates[0], dates[-1]) if c.get("media_file_id")]
    all_workers = db.get_active_workers()
    last_groups = db.get_workers_last_groups()

    wb = Workbook()
    ws = wb.active
    ws.title = _month_label(year, mon)
    _write_matrix_sheet(
        ws, f"{i18n.MONTHLY_EXCEL_TITLE}: {_month_label(year, mon)}",
        dates, checkins, all_workers, last_groups, tz,
    )

    ws2 = wb.create_sheet(i18n.EXCEL_SHEET_CHECKINS)
    _write_detail_sheet(ws2, checkins, tz)

    now_str = datetime.now(tz).strftime("%Y%m%d_%H%M%S")
    filename = f"{title}_{year:04d}-{mon:02d}_{now_str}.xlsx"
    filepath = os.path.join(config.EXPORTS_DIR, filename)
    wb.save(filepath)
    return filepath


def generate_alltime_export(title: str = "alltime") -> str:
    """Create an all-time attendance .xlsx: one weekly-style matrix sheet per
    month of history.

    All-time can span tens of thousands of check-ins, so this uses a lightweight
    (no-join) fetch and shows only the month matrices — no giant per-check-in
    sheet (use "Экспорт всех данных" for the raw dump). This keeps it fast enough
    to finish inside the serverless time limit.
    """
    os.makedirs(config.EXPORTS_DIR, exist_ok=True)
    tz = ZoneInfo(config.TIMEZONE)
    today = datetime.now(tz).date()

    # Lightweight fetch: only the columns the matrix needs (no workers/groups join).
    raw = db.get_all_checkins(columns="user_id, group_id, date, timestamp, media_file_id")
    checkins = [c for c in raw
                if c.get("media_file_id") and c.get("group_id") not in config.EXCLUDED_GROUP_IDS]
    all_workers = db.get_active_workers()

    # Group names come from the small groups table (excluded groups already dropped there).
    groups_by_id = {g["group_id"]: (g.get("group_name") or "—") for g in db.get_all_groups()}
    last_groups = {}
    for c in checkins:  # ascending time → last write wins
        if c["group_id"] in groups_by_id:
            last_groups[c["user_id"]] = groups_by_id[c["group_id"]]

    # Months that actually have data (fall back to the current month if none).
    months = sorted({c["date"][:7] for c in checkins}) or [today.strftime("%Y-%m")]

    wb = Workbook()
    for i, ym in enumerate(months):
        year, mon = int(ym[:4]), int(ym[5:7])
        dates = _month_day_list(year, mon, today)
        month_checkins = [c for c in checkins if c["date"][:7] == ym]
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = _month_label(year, mon)
        _write_matrix_sheet(
            ws, f"{i18n.ALLTIME_EXCEL_TITLE}: {_month_label(year, mon)}",
            dates, month_checkins, all_workers, last_groups, tz,
        )

    now_str = datetime.now(tz).strftime("%Y%m%d_%H%M%S")
    filename = f"{title}_{now_str}.xlsx"
    filepath = os.path.join(config.EXPORTS_DIR, filename)
    wb.save(filepath)
    return filepath
